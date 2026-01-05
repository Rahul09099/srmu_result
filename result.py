# import requests
# USER_ID = "202410101200106"
# PASSWORD = "123"



ROLL_NUMBERS=[
    "202210101200008",
    "202210101200006",
    "202210101200017",
    "202210101200012",
    "202210101200015",
    "202210101200007",
    "202210101200013",
    "202210101200009",
    "202210101200002",
    "202210101200019",
    "202210101200010",
    "202210101200005",
    "202210101200022",
    "202210101200001",
    "202210101200018",
    "202210101200004",
    "202210101200011",
    "202210101200020",
    "202210101200021",
    "202210101200003",
    "202210101200016",
    "202210101200014",
    "202210101200023"
]


# ================= CONFIG =================
# ROLL_NUMBERS= [
#     "202410101200001",
#     "202410101200003",
#     "202410101200004",
#     "202410101200005",
#     "202410101200006",
#     "202410101200007",
#     "202410101200008",
#     "202410101200009",
#     "202410101200010",
#     "202410101200011",
#     "202410101200012",
#     "202410101200013",
#     "202410101200014",
#     "202410101200015",
#     "202410101200016",
#     "202410101200017",
#     "202410101200018",
#     "202410101200019",
#     "202410101200020",
#     "202410101200021",
#     "202410101200022",
#     "202410101200023",
#     "202410101200024",
#     "202410101200025",
#     "202410101200028",
#     "202410101200029",
#     "202410101200030",
#     "202410101200031",
#     "202410101200032",
#     "202410101200033",
#     "202410101200034",
#     "202410101200035",
#     "202410101200036",
#     "202410101200037",
#     "202410101200038",
#     "202410101200039",
#     "202410101200040",
#     "202410101200041",
#     "202410101200042",
#     "202410101200043",
#     "202410101200044",
#     "202410101200045",
#     "202410101200047",
#     "202410101200048",
#     "202410101200050",
#     "202410101200051",
#     "202410101200052",
#     "202410101200053",
#     "202410101200054",
#     "202410101200055",
#     "202410101200056",
#     "202410101200057",
#     "202410101200058",
#     "202410101200059",
#     "202410101200060",
#     "202410101200061",
#     "202410101200062",
#     "202410101200063",
#     "202410101200064",
#     "202410101200065",
#     "202410101200066",
#     "202410101200067",
#     "202410101200068",
#     "202410101200069",
#     "202410101200070",
#     "202410101200071",
#     "202410101200072",
#     "202410101200073",
#     "202410101200074",
#     "202410101200075",
#     "202410101200076",
#     "202410101200077",
#     "202410101200078",
#     "202410101200079",
#     "202410101200080",
#     "202410101200081",
#     "202410101200082",
#     "202410101200083",
#     "202410101200084",
#     "202410101200085",
#     "202410101200086",
#     "202410101200087",
#     "202410101200088",
#     "202410101200089",
#     "202410101200091",
#     "202410101200092",
#     "202410101200093",
#     "202410101200094",
#     "202410101200095",
#     "202410101200096",
#     "202410101200097",
#     "202410101200098",
#     "202410101200099",
#     "202410101200100",
#     "202410101200101",
#     "202410101200102",
#     "202410101200103",
#     "202410101200104",
#     "202410101200105",
#     "202410101200106",
#     "202410101200107",
#     "202410101200108",
#     "202410101200109",
#     "202410101200110",
#     "202410101200111",
#     "202410101200112",
#     "202410101200113",
#     "202410101200115",
#     "202410101200116",
#     "202410101200117",
#     "202410101200119",
#     "202410101200120",
#     "202410101200121",
#     "202410101200122",
#     "202410101200123",
#     "202410101200124",
#     "202410101200125",
#     "202410101200127",
#     "202410101200128",
#     "202410101200129",
#     "202410101200130",
#     "202410101200132",
#     "202410101200133",
#     "202410101200134",
#     "202410101200135",
#     "202410101200136",
#     "202410101200137",
#     "202410101200138",
#     "202410101200139",
#     "202410101200140",
#     "202410101200141",
#     "202410101200142",
#     "202410101200143",
#     "202410101200144",
#     "202410101200145",
#     "202410101200146",
#     "202410101200147",
#     "202410101200149",
#     "202410101200151",
#     "202410101200152",
#     "202410101200153",
#     "202410101200154",
#     "202410101200156",
#     "202410101200157",
#     "202410101200158",
# ]

from playwright.sync_api import sync_playwright
import requests
import pdfplumber
import re
import os
import time
from openpyxl import Workbook, load_workbook

# ================= CONFIG =================
# ROLL_NUMBERS = [
#     "202410101200001",
#     "202410101200003",
#     "202410101200004",
#     # ... keep your full list
#     "202410101200158",
# ]

PASSWORD = "123"
TERM = "2501"

DOWNLOAD_DIR = "downloads"
os.makedirs(DOWNLOAD_DIR, exist_ok=True)

EXCEL_FILE = "Transcript_Data.xlsx"
SUCCESS_LOG = "success_rolls.txt"
FAILED_LOG = "failed_rolls.txt"

DELAY_BETWEEN_STUDENTS = 3  # seconds
# =========================================


# ========== LOG HELPERS ==========
def log_success(roll):
    with open(SUCCESS_LOG, "a") as f:
        f.write(f"{roll}\n")


def log_failure(roll, reason):
    with open(FAILED_LOG, "a") as f:
        f.write(f"{roll} | {reason}\n")


# ========== PDF HELPERS ==========
def read_pdf_text(pdf_path):
    text = ""
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            if page.extract_text():
                text += page.extract_text() + "\n"
    return text


def extract_from_text(text):
    # Student name
    name_match = re.search(r"Student's Name:\s*([A-Za-z ]+)", text)
    name = name_match.group(1).strip() if name_match else "Unknown"

    # Subjects
    subjects = []
    subject_pattern = re.compile(
        r"(BCS\d{4}|BHU\d{4})\s+(.+?)\s+(A\+?|B\+?|C\+?|D|F)\s+\d+\s+\d+\s+\d+"
    )

    for m in subject_pattern.finditer(text):
        subjects.append({
            "name": m.group(2).strip(),
            "grade": m.group(3)
        })

    # SGPA & CGPA
    sgpa_cgpa = re.search(
        r"SGPA\s+CGPA.*?\n([\d.]+)\s+([\d.]+)",
        text,
        re.DOTALL
    )

    sgpa = sgpa_cgpa.group(1) if sgpa_cgpa else "NA"
    cgpa = sgpa_cgpa.group(2) if sgpa_cgpa else "NA"

    return name, subjects, sgpa, cgpa


# ========== EXCEL INIT ==========
COURSE_NAMES_FIXED = None

if os.path.exists(EXCEL_FILE):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.title = "Transcript Data"


# ========== PLAYWRIGHT ==========
with sync_playwright() as p:
    browser = p.chromium.launch(headless=False)

    for USER_ID in ROLL_NUMBERS:
        print(f"\n🔹 Processing Roll No: {USER_ID}")
        context = browser.new_context()
        page = context.new_page()

        try:
            # 1️⃣ LOGIN
            page.goto("https://eyojan.srmu.ac.in/psc/ps/?cmd=login&languageCd=ENG")
            page.fill("input[name='userid']", USER_ID)
            page.fill("input[name='pwd']", PASSWORD)

            print("👉 Complete captcha / OTP manually if asked")
            page.click("input[type='submit']")
            page.wait_for_load_state("networkidle", timeout=120000)

            # 2️⃣ OPEN TRANSCRIPT PAGE
            page.goto(
                "https://eyojan.srmu.ac.in/psc/ps/EMPLOYEE/SA/c/SRMU_MENU.SRMU_STD_UN_TR_CMP.GBL"
            )
            page.wait_for_load_state("networkidle")

            # 3️⃣ SET TERM
            page.wait_for_selector("#SRM_EMP_TERM_WK_STRM", timeout=60000)
            term_input = page.locator("#SRM_EMP_TERM_WK_STRM")
            term_input.fill("")
            term_input.type(TERM)
            term_input.press("Tab")
            page.wait_for_load_state("networkidle")

            # 4️⃣ GENERATE PDF
            with page.expect_popup() as popup_info:
                page.click("#SRMU_TR_INDV_WK_BUTTON")

            pdf_page = popup_info.value
            pdf_page.wait_for_load_state("load")
            pdf_url = pdf_page.url

            # 5️⃣ DOWNLOAD PDF
            cookies = context.cookies()
            cookie_dict = {c["name"]: c["value"] for c in cookies}

            pdf_response = requests.get(pdf_url, cookies=cookie_dict)

            if len(pdf_response.content) < 10000:
                log_failure(USER_ID, "Invalid or empty PDF")
                print(f"⚠️ Invalid PDF for {USER_ID}")
                continue

            temp_pdf = os.path.join(DOWNLOAD_DIR, f"{USER_ID}.pdf")
            with open(temp_pdf, "wb") as f:
                f.write(pdf_response.content)

            pdf_page.close()

            # 6️⃣ EXTRACT DATA
            text = read_pdf_text(temp_pdf)
            name, subjects, sgpa, cgpa = extract_from_text(text)

            if not subjects:
                log_failure(USER_ID, "No subjects found in PDF")
                print(f"⚠️ No data in PDF for {USER_ID}")
                continue

            # 7️⃣ FIX COURSE STRUCTURE (FIRST SUCCESS ONLY)
            if COURSE_NAMES_FIXED is None:
                COURSE_NAMES_FIXED = [s["name"] for s in subjects]
                headers = ["Name"] + COURSE_NAMES_FIXED + ["SGPA", "CGPA"]
                if ws.max_row == 1:
                    ws.append(headers)
                    wb.save(EXCEL_FILE)

            # 8️⃣ PREPARE ROW
            grade_map = {s["name"]: s["grade"] for s in subjects}

            row = [name]
            for course_name in COURSE_NAMES_FIXED:
                row.append(grade_map.get(course_name, ""))

            row.extend([sgpa, cgpa])
            ws.append(row)
            wb.save(EXCEL_FILE)

            log_success(USER_ID)
            print(f"💾 Data saved for {USER_ID}")

        except Exception as e:
            print(f"❌ Error for {USER_ID}: {e}")
            log_failure(USER_ID, str(e))

        finally:
            context.close()
            time.sleep(DELAY_BETWEEN_STUDENTS)

    browser.close()


print("\n✅ SCRIPT COMPLETED")
print("📄 Excel:", EXCEL_FILE)
print("📄 Success log:", SUCCESS_LOG)
print("📄 Failed log:", FAILED_LOG)
